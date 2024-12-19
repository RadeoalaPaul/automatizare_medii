from PyQt6 import uic
from PyQt6 import QtWidgets
from PyQt6.QtWidgets import QApplication, QLabel, QDialog, QPushButton, QTextEdit
import sys
import pathlib
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
import re
import os

tip_activitate = 0
rand_disciplina = 0
indice_max = 0
label_status = None

def activitate(tip):
   global tip_activitate
   tip_activitate = tip

def disciplina(numar):
   global rand_disciplina
   rand_disciplina = numar
   stack.setCurrentIndex(stack.currentIndex()+1)

def actualizare(a, d, n): #a - activitate, d - tip disciplina

   global letterS, letterM, indice_max

   actualizat = False
   randuri = []
   coloane = []
   medii = []
   s = 0
   workbook = load_workbook("mediii.xlsx")
   #workbook = load_workbook("medii_copie.xlsx") #pt test
   sheet = workbook.active
   for col in sheet.iter_rows(min_col=1,max_col=sheet.max_column,min_row=1,max_row=sheet.max_row):
      for cell in col:
         if(cell.value == "Laborator" or cell.value == "Seminar"):
            randuri.append(cell.row)
         
   # d+a linia activitatii
   #scriere functional
   for col in sheet.iter_cols(min_col=3,max_col=sheet.max_column+1,min_row=d+a,max_row=d+a):
      for cell in col:
         if(cell.value == None and not actualizat):
            sheet[str(get_column_letter(cell.col_idx))+str(d+a)] = int(n)
            actualizat = True

   #calcul medii functional
   for i in range(1,23):
      if randuri.count(i) >= 1:
         s = 0
         c = 0
         for col in sheet.iter_cols(min_col=3,max_col=sheet.max_column,min_row=i,max_row=i):
            for cell in col:
               if(cell.value!=None):
                  s+=cell.value
                  c+=1
         if(c!=0):
            medii.append(s/c)
         else:
            medii.append(None)
   for rand, media in zip(randuri, medii):
      sheet.cell(row=rand, column=1, value=media)
      if(media!=None):
         if(media>=5):
            sheet.cell(row=rand, column=1, value=media).fill = PatternFill("solid", fgColor="60fb27")
         else:
            sheet.cell(row=rand, column=1, value=media).fill = PatternFill("solid", fgColor="fb390f")

   workbook.save("mediii.xlsx")
   
def trimite_email(text):
   email = re.match(r"^\S+@\S+\.\S+$",text) #functional
   if(email==None):
      label_status.setText("Status: E-mail invalid!")
   else:
      label_status.setText("Status: Fisier trimis cu succes!")

      sender_email = "radeoalapaul34@gmail.com"
      receiver_email = text
      password = "took gvgb gzoo cdie"
      server = "smtp.gmail.com"
      subject = "Medii actualizate facultate"
      body = "Medii actualizate"
      
      file_path = Path.cwd() / 'mediii.xlsx'
      
      msg = EmailMessage()
      msg["From"] = sender_email
      msg["To"] = receiver_email
      msg["Subject"] = subject
      msg.set_content(body)

      try:
         with open(file_path, "rb") as file:
            file_data = file.read()
            #file_name = file.name.split("/")[-1] # Numele fișierului atașat
            file_name = os.path.basename(file_path)
         msg.add_attachment(file_data, maintype="application",
         subtype="octet-stream", filename=file_name)

      except FileNotFoundError:
         print(f"Fișierul {file_path} nu a fost găsit.")
         exit()
         
      try:
         with smtplib.SMTP_SSL(server, 465) as smtp:
            smtp.login(sender_email, password)
            smtp.send_message(msg)
            print("E-mail trimis cu succes!")
      except Exception as e:
         print(f"Eroare: {e}")

class MainWindow(QDialog):
   def __init__(self):
      global buton_seminar_widget, buton_laborator_widget
      super().__init__()

      uic.loadUi("medii_activitate.ui", self)

      buton_seminar_widget = self.findChild(QPushButton,"buton_seminar")
      buton_laborator_widget = self.findChild(QPushButton,"buton_laborator")

      buton_seminar_widget.clicked.connect(self.activitate_seminar)
      buton_laborator_widget.clicked.connect(self.activitate_laborator)

   def activitate_seminar(self):
       activitate(1)
       stack.setCurrentIndex(stack.currentIndex()+1)
   def activitate_laborator(self):
       activitate(2)
       stack.setCurrentIndex(stack.currentIndex()+1)

class Afisare_Discipline(QDialog):
   def __init__(self):
      global buton_inapoi_widget
      global buton_d1,buton_d2,buton_d3,buton_d4,buton_d5,buton_d6,buton_d7
      super().__init__()

      uic.loadUi("medii_disciplina.ui",self)

      buton_inapoi_widget = self.findChild(QPushButton,"buton_inapoi")

      buton_inapoi_widget.clicked.connect(self.inapoi)

      buton_d1 = self.findChild(QPushButton,"buton_d1")
      buton_d2 = self.findChild(QPushButton,"buton_d2")
      buton_d3 = self.findChild(QPushButton,"buton_d3")
      buton_d4 = self.findChild(QPushButton,"buton_d4")
      buton_d5 = self.findChild(QPushButton,"buton_d5")
      buton_d6 = self.findChild(QPushButton,"buton_d6")
      buton_d7 = self.findChild(QPushButton,"buton_d7")
   
      buton_d1.clicked.connect(self.d1)
      buton_d2.clicked.connect(self.d2)
      buton_d3.clicked.connect(self.d3)
      buton_d4.clicked.connect(self.d4)
      buton_d5.clicked.connect(self.d5)
      buton_d6.clicked.connect(self.d6)
      buton_d7.clicked.connect(self.d7)

   def d1(self):
      disciplina(2) #setam variabila echivalenta cu randul titlului materiei respective
   def d2(self):
      disciplina(5)
   def d3(self):
      disciplina(8)
   def d4(self):
      disciplina(11)
   def d5(self):
      disciplina(14)
   def d6(self):
      disciplina(17)
   def d7(self):
      disciplina(20) 

   def inapoi(self):
      stack.setCurrentIndex(stack.currentIndex()-1)

class Afisare_Nota(QDialog):
   def __init__(self):
      global buton_proces_widget, buton_inapoi_widget, text_email, text_nota, label_status
      super().__init__()

      uic.loadUi("medii_nota.ui", self)

      buton_inapoi_widget = self.findChild(QPushButton, "buton_inapoi")
      buton_proces_widget = self.findChild(QPushButton, "buton_proces")
      label_status = self.findChild(QLabel, "label_status")
      text_nota = self.findChild(QTextEdit,"text_nota")
      text_email = self.findChild(QTextEdit,"text_email")

      buton_proces_widget.clicked.connect(self.proces)
      buton_inapoi_widget.clicked.connect(self.inapoi)

   def proces(self):
      actualizare(tip_activitate,rand_disciplina,text_nota.toPlainText())
      trimite_email(text_email.toPlainText())

   def inapoi(self):
      stack.setCurrentIndex(stack.currentIndex()-1)

app = QApplication(sys.argv)
stack = QtWidgets.QStackedWidget()
window = MainWindow()
afisare_discipline = Afisare_Discipline()
afisare_nota = Afisare_Nota()
stack.addWidget(window)
stack.addWidget(afisare_discipline)
stack.addWidget(afisare_nota)
stack.setFixedHeight(600)
stack.setFixedWidth(305)
stack.show()
sys.exit(app.exec())